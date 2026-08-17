"""
bom_extractor.py — itemized FG->SFG->RM Bill of Materials extraction from
Ramco's monthly ledger exports.

rm_ledger.py already sums a *total* RM cost per product from the FG sheet.
This module extracts the individual *component rows* instead — the actual
Bill of Materials shown in the "Material Costing" tab.

Ledger schema (confirmed against the real May'26 FG/SFG sheets — both share
the same 28 columns):
  Bom Code/PS.No   — the parent product being built (FG code in the FG
                      sheet, SFG code in the SFG sheet)
  Issue Item       — the component actually consumed (RM or SFG), followed
                      immediately by ITS OWN Item Desc / Uom / Qty columns
  Std Wh Code      — the issued component's own warehouse; 'CBESFG' means
                      the component is itself an SFG code (needs expanding
                      via a second lookup against the SFG sheet's own
                      Bom Code/PS.No column) — anything else is treated as
                      a raw material.
  Rate / Total <Month> — unit rate and the ledger's own pre-computed cost
                      for that line. Total <Month> is used as the line's
                      cost directly (NOT qty*rate) because some items (e.g.
                      thread, priced per spool) carry a hidden unit-
                      conversion factor the simple Qty/Rate columns don't
                      reflect on their own — Ramco's own ledger already
                      applies it in Total <Month>, so that's the only
                      trustworthy cost figure. (This is the same 18x
                      spool-pricing trap this project already hit and fixed
                      once for the mockup's hand-extracted Comfort 6" BOM.)

Column name caveat: "Item Desc", "Uom", and "Qty" each appear TWICE in the
header row (once for the "Item" column, once for "Issue Item") — headers
.index() would silently grab the wrong (first) occurrence for the second
copy, so those three are addressed positionally, immediately following
"Issue Item", not by name.
"""
from __future__ import annotations
from typing import Iterable
import sys
import time
import openpyxl

from rm_ledger import _find_sheet, _find_total_col

# How often (in rows scanned) to refresh the live progress line — frequent
# enough to feel alive on a multi-hundred-thousand-row sheet, infrequent
# enough that the print() calls themselves don't slow the scan down.
_PROGRESS_EVERY = 50_000


def _report(prefix: str, count: int, t0: float, done: bool = False):
    elapsed = time.time() - t0
    rate = count / elapsed if elapsed > 0 else 0
    end = '\n' if done else ''
    msg = f"\r  {prefix}: {count:,} rows scanned in {elapsed:,.0f}s ({rate:,.0f} rows/s)"
    print(msg + (' — done' if done else ''), end=end, flush=True)


def _safe_float(v):
    # Some ledgers (e.g. the HCF workbooks) have real rows with a literal
    # '#N/A' (or other Excel error text) in Rate/Qty/Total instead of a
    # number — treat as 0 rather than crashing the whole extraction on
    # one bad cell.
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def extract_boms(item_codes: Iterable[str], fg_path: str, sfg_path: str, progress: bool = False) -> dict:
    """
    Extract itemized BOMs for every item_code found in the FG ledger.

    Two total passes regardless of how many item_codes are requested (one
    over the FG sheet, one over the SFG sheet) — same efficient
    filter-while-streaming approach as rm_ledger.sum_rm_cost, so a batch of
    50+ SKUs costs the same as extracting one.

    progress=True prints a live, self-overwriting "N rows scanned" line
    (refresh_bom_data.py's terminal run enables this) — these ledgers run
    to the hundreds of thousands of rows and each sheet can take several
    minutes, so silence for that long is easy to mistake for a hang.

    Returns {item_code: [component, ...]}. Each component:
      {code, desc, wh ('RM'|'SFG'), qty, uom, rate, cost, children}
    'children' is only present (a list, possibly empty) when wh == 'SFG'.
    item_codes with no matching FG-sheet rows are simply absent from the
    result — never fabricated as an empty list, so the caller can tell
    "not found in this month's ledger" apart from "found, zero lines".
    """
    wanted = set(item_codes)
    boms: dict[str, list[dict]] = {}
    sfg_codes_needed: set[str] = set()

    # keep_links=False: skips eagerly parsing the external-links XML, which
    # openpyxl does even in read_only mode and which can raise MemoryError
    # on some of these ledger exports on a memory-constrained machine —
    # this function only ever reads cell values, never external references.
    wb = openpyxl.load_workbook(fg_path, read_only=True, data_only=True, keep_links=False)
    try:
        ws = wb[_find_sheet(wb, 'FG')]
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or '').strip() for c in next(rows)]
        bc_i = headers.index('Bom Code/PS.No')
        ii_i = headers.index('Issue Item')
        rate_i = headers.index('Rate')
        tot_i = _find_total_col(headers)
        idesc_i, iuom_i, iqty_i, wh_i = ii_i + 1, ii_i + 2, ii_i + 3, ii_i + 4

        t0 = time.time()
        scanned = 0
        for row in rows:
            scanned += 1
            if progress and scanned % _PROGRESS_EVERY == 0:
                _report('FG sheet', scanned, t0)
            if not row:
                continue
            parent = row[bc_i]
            if parent not in wanted:
                continue
            code = row[ii_i]
            if not code:
                continue
            wh = 'SFG' if str(row[wh_i] or '').strip().upper() == 'CBESFG' else 'RM'
            comp = {
                'code': str(code).strip(),
                'desc': str(row[idesc_i] or '').strip(),
                'wh': wh,
                'qty': _safe_float(row[iqty_i]),
                'uom': str(row[iuom_i] or '').strip(),
                'rate': _safe_float(row[rate_i]),
                'cost': round(_safe_float(row[tot_i]), 4),
            }
            boms.setdefault(parent, []).append(comp)
            if wh == 'SFG':
                sfg_codes_needed.add(comp['code'])
        if progress:
            _report('FG sheet', scanned, t0, done=True)
    finally:
        wb.close()

    children_by_sfg: dict[str, list[dict]] = {}
    if sfg_codes_needed:
        wb2 = openpyxl.load_workbook(sfg_path, read_only=True, data_only=True, keep_links=False)
        try:
            ws2 = wb2[_find_sheet(wb2, 'SFG')]
            rows2 = ws2.iter_rows(values_only=True)
            headers2 = [str(c or '').strip() for c in next(rows2)]
            bc_i2 = headers2.index('Bom Code/PS.No')
            ii_i2 = headers2.index('Issue Item')
            rate_i2 = headers2.index('Rate')
            tot_i2 = _find_total_col(headers2)
            idesc_i2, iuom_i2, iqty_i2 = ii_i2 + 1, ii_i2 + 2, ii_i2 + 3

            t0 = time.time()
            scanned2 = 0
            for row in rows2:
                scanned2 += 1
                if progress and scanned2 % _PROGRESS_EVERY == 0:
                    _report('SFG sheet', scanned2, t0)
                if not row:
                    continue
                parent = row[bc_i2]
                if parent not in sfg_codes_needed:
                    continue
                code = row[ii_i2]
                if not code:
                    continue
                children_by_sfg.setdefault(parent, []).append({
                    'code': str(code).strip(),
                    'desc': str(row[idesc_i2] or '').strip(),
                    'qty': _safe_float(row[iqty_i2]),
                    'uom': str(row[iuom_i2] or '').strip(),
                    'rate': _safe_float(row[rate_i2]),
                    'cost': round(_safe_float(row[tot_i2]), 4),
                })
            if progress:
                _report('SFG sheet', scanned2, t0, done=True)
        finally:
            wb2.close()

    for comps in boms.values():
        for c in comps:
            if c['wh'] == 'SFG':
                c['children'] = children_by_sfg.get(c['code'], [])

    return boms
