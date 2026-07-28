"""
rm_ledger.py — parse Ramco's monthly FG/SFG ledger exports for per-SKU RM cost.

These files already contain the fully-resolved, per-component costed BOM —
SFG rollups and unit-of-measure conversions are already done by Ramco itself
(see "Conv Factor" / "Final Qty" columns). This module just streams the
sheet and sums the "Total <month>" column per Bom Code/PS.No (= item_code).
It never loads the whole (80-250MB) file into memory.

Column-name caveat: the "Total" column is named inconsistently across
months (e.g. "Total May'26", "Total Mar'26", or plain "Total" for Feb'26)
— _find_total_col() matches any header starting with "Total" rather than
a fixed string, so this survives that drift.
"""
from __future__ import annotations
from typing import Iterable
import openpyxl


def _find_total_col(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if h.strip().lower().startswith('total'):
            return i
    raise ValueError(f"No 'Total' column found in headers: {headers}")


def _find_sheet(wb, keyword: str) -> str:
    if keyword in wb.sheetnames:
        return keyword
    matches = [s for s in wb.sheetnames if keyword.upper() in s.upper()]
    if not matches:
        raise ValueError(f"No sheet matching '{keyword}' in {wb.sheetnames}")
    return matches[0]


def sum_rm_cost(filepath: str, item_codes: Iterable[str], sheet_keyword: str = 'FG') -> dict:
    """
    Sum the Total-<month> column per Bom Code/PS.No for the given item_codes.
    Returns {item_code: {'rm_cost': float, 'line_count': int}}.
    """
    wanted = set(item_codes)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet_name = _find_sheet(wb, sheet_keyword)
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or '').strip() for c in next(rows)]
        bc_i = headers.index('Bom Code/PS.No')
        tot_i = _find_total_col(headers)

        totals: dict[str, float] = {}
        counts: dict[str, int] = {}
        for row in rows:
            if not row:
                continue
            code = row[bc_i]
            if code in wanted:
                totals[code] = totals.get(code, 0.0) + float(row[tot_i] or 0)
                counts[code] = counts.get(code, 0) + 1
        return {c: {'rm_cost': round(totals[c], 4), 'line_count': counts[c]} for c in totals}
    finally:
        wb.close()
