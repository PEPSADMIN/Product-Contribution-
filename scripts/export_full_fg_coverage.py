"""
export_full_fg_coverage.py — the complete, unrestricted picture: every
FG (finished-goods) AND SFG (semi-finished-goods / sub-assembly) code
found ANYWHERE — Item Master's full catalog, each of the 4 monthly FG
and SFG ledgers, and each month's "Foam Item Codes - Rate" workbook —
with a per-month/per-source Yes/No presence flag, so the user can hand
this to their team and see exactly what's available and what's missing.
Nothing is filtered down to only the 51/72 SKUs this tool calculates
margins for; two full sheets ("FG Coverage" and "SFG Coverage") cover
both levels of the BOM tree.

The Foam Item Codes files are SFG-level (Std Wh Code = 'CBESFG' on the
Foam-KG/Foam-MTR/HYP-Block sheets — component/sub-assembly foam, not
top-level finished products), so their codes are cross-referenced on
BOTH sheets: on FG Coverage because the user reviews that file for foam
costing regardless of level, and on SFG Coverage because that's their
actual classification.

RM Cost / itemized BOM data only shows up for the SKUs this tool has
extracted (the 51 in sku_master.py) — extending that to every code in
these sheets would mean itemized BOM extraction for tens of thousands
of codes, a separate, much bigger job this script doesn't attempt. For
SFG codes, "used in tool's BOM data?" instead checks whether that SFG
code appears as a component anywhere across the 51 tracked SKUs'
already-extracted BOMs.
"""
import os
import sys
import time

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from rm_ledger import _find_sheet
from item_master import get_index
from sku_master import SKUS
import costing_store
import bom_store

RM_ROOT = r"C:\Users\ADMIN\Downloads\Product Contribution\Accounts W - 27.07.2026\RM"
OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "validation_exports", "Full_FG_Coverage_Validation.xlsx")

MONTH_LABELS = ["Feb'26", "Mar'26", "April'26", "May'26"]

FG_FILES = [
    r"1 - Feb'26\2 - FG Feb'26.xlsx",
    r"2 - Mar'26\1 - FG Mar'26 dt 04-04-2026.xlsx",
    r"3 - April'26\2 - FG April'26 dt 22-05-2026.xlsx",
    r"4 - May'26\2 - FG MAY'26.xlsx",
]

SFG_FILES = [
    r"1 - Feb'26\1 - SFG Feb'26.xlsx",
    r"2 - Mar'26\2 - SFG Mar'26 dt 03-04-2026.xlsx",
    r"3 - April'26\1 - SFG April'26 dt 22-05-2026.xlsx",
    r"4 - May'26\1 - SFG MAY'26.xlsx",
]

# Filenames aren't consistent month to month (some have a leading "4 - ",
# April's file is even internally labelled "Mar'26" despite living in the
# April folder) — hardcoded per month, same as FG_FILES/SFG_FILES above.
FOAM_FILES = [
    r"1 - Feb'26\Foam Item Codes - Rate - Feb'26 Mail.xlsx",
    r"2 - Mar'26\4 - Foam Item Codes - Rate - Mar'26 SCR 31-03-2026.xlsx",
    r"3 - April'26\4 - Foam Item Codes - Rate - Mar'26 Mail dt 13-05-2026.xlsx",
    r"4 - May'26\4 - Foam Item Codes - Rate - May'26 Mail.xlsx",
]

# Column names actually seen across these workbooks' various sheets
# (Foam-KG/Foam-MTR/HYP-Block use "Issue Item" or "Item Code"; the
# STD-*/PL-* cost-breakdown sheets use "ERP Item Code" or "Item Code").
_CODE_COL_NAMES = {'issue item', 'item code', 'erp item code'}
_DESC_COL_NAMES = {'item desc', 'item discription', 'item description'}

_PROGRESS_EVERY = 50_000


def scan_ledger_codes(path: str, sheet_type: str) -> dict:
    """Every distinct (Bom Code/PS.No -> PS Desc.) found in one month's
    FG or SFG sheet — both share the same 28-column schema."""
    codes = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_find_sheet(wb, sheet_type)]
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or '').strip() for c in next(rows)]
        bc_i = headers.index('Bom Code/PS.No')
        desc_i = headers.index('PS Desc.')

        t0 = time.time()
        scanned = 0
        for row in rows:
            scanned += 1
            if scanned % _PROGRESS_EVERY == 0:
                elapsed = time.time() - t0
                rate = scanned / elapsed if elapsed else 0
                print(f"\r    {scanned:,} rows scanned in {elapsed:,.0f}s ({rate:,.0f} rows/s), "
                      f"{len(codes):,} distinct codes so far", end='', flush=True)
            if not row:
                continue
            code = row[bc_i]
            if not code:
                continue
            code = str(code).strip()
            desc = str(row[desc_i] or '').strip()
            if code not in codes or (desc and not codes[code]):
                codes[code] = desc
        print(f"\r    {scanned:,} rows scanned in {time.time()-t0:,.0f}s — done, "
              f"{len(codes):,} distinct codes" + " " * 10)
    finally:
        wb.close()
    return codes


def scan_foam_codes(path: str) -> dict:
    """
    Every distinct item code found across ALL sheets of a "Foam Item
    Codes - Rate" workbook — scans every sheet generically (rather than
    hardcoding sheet names, which change month to month) by looking for a
    header row containing one of _CODE_COL_NAMES within the first few
    rows of each sheet (the STD-*/PL-* sheets have a title row before
    their real header).
    """
    codes = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            headers = None
            for _ in range(3):
                try:
                    row = next(rows)
                except StopIteration:
                    row = None
                    break
                if row and any(str(c or '').strip().lower() in _CODE_COL_NAMES for c in row):
                    headers = [str(c or '').strip().lower() for c in row]
                    break
            if not headers:
                continue
            code_i = next((i for i, h in enumerate(headers) if h in _CODE_COL_NAMES), None)
            desc_i = next((i for i, h in enumerate(headers) if h in _DESC_COL_NAMES), None)
            if code_i is None:
                continue
            for row in rows:
                if not row or code_i >= len(row):
                    continue
                code = row[code_i]
                if not code:
                    continue
                code = str(code).strip()
                desc = str(row[desc_i] or '').strip() if desc_i is not None and desc_i < len(row) else ''
                if code not in codes or (desc and not codes[code]):
                    codes[code] = desc
    finally:
        wb.close()
    return codes


def build_coverage_sheet(wb, sheet_title, code_label, all_codes, item_master, item_master_label,
                          per_month_ledger, per_month_foam, tracked_col_headers, tracked_row_fn):
    """
    Shared sheet-builder for both FG and SFG coverage — same column
    layout, sorting, and highlighting logic either way; only the data
    sources and the "tracked" columns' meaning differ.

    tracked_row_fn(code) -> list of values matching tracked_col_headers,
    plus the first of those values must be "Yes"/"No" (used for sorting
    and the green-row highlight).
    """
    rows = []
    for code in all_codes:
        desc = (item_master.get(code)
                or next((per_month_ledger[lbl].get(code) for lbl in MONTH_LABELS if per_month_ledger[lbl].get(code)), '')
                or next((per_month_foam[lbl].get(code) for lbl in MONTH_LABELS if per_month_foam[lbl].get(code)), ''))
        in_foam_any = any(code in per_month_foam[lbl] for lbl in MONTH_LABELS)
        row = [code, desc, "Yes" if code in item_master else "No"]
        row += ["Yes" if code in per_month_ledger[lbl] else "No" for lbl in MONTH_LABELS]
        row += ["Yes" if in_foam_any else "No"]
        row += ["Yes" if code in per_month_foam[lbl] else "No" for lbl in MONTH_LABELS]
        row += tracked_row_fn(code)
        rows.append(row)

    # Column layout (0-indexed): 0=code 1=desc 2=in_item_master
    # 3..6 = 4 ledger months, 7 = in_foam_any, 8..11 = 4 foam months,
    # 12.. = tracked_col_headers
    TRACKED_COL, FOAM_ANY_COL = 12, 7

    def sort_key(r):
        tracked_flag = r[TRACKED_COL] != "Yes"
        missing_somewhere = "No" in r[2:7] or r[FOAM_ANY_COL] == "No"
        return (tracked_flag, not missing_somewhere, r[0])
    rows.sort(key=sort_key)

    ws = wb.create_sheet(sheet_title) if sheet_title not in wb.sheetnames else wb[sheet_title]

    headers = ([code_label, "Description", f"In {item_master_label}?"]
               + [f"In {lbl} Ledger?" for lbl in MONTH_LABELS]
               + ["In Foam Codes File?"]
               + [f"In {lbl} Foam File?" for lbl in MONTH_LABELS]
               + tracked_col_headers)
    ws.append([f"{sheet_title} — {item_master_label} + all 4 monthly ledgers + all 4 Foam Item Codes files ({len(rows):,} total)"])
    ws.append([])
    ws.append(headers)
    header_row = 3
    for r in rows:
        ws.append(r)

    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = ([26, 46, 13] + [13] * len(MONTH_LABELS) + [15] + [13] * len(MONTH_LABELS)
              + [18] * len(tracked_col_headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    gap_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    tracked_fill = PatternFill(start_color="DCEFE3", end_color="DCEFE3", fill_type="solid")
    n_cols = len(headers)
    tracked_excel_col = TRACKED_COL + 1  # 1-indexed Excel column
    for r in range(header_row + 1, header_row + 1 + len(rows)):
        gap_check_vals = [ws.cell(row=r, column=c).value for c in range(3, 3 + 1 + len(MONTH_LABELS))]
        gap_check_vals.append(ws.cell(row=r, column=FOAM_ANY_COL + 1).value)
        if ws.cell(row=r, column=tracked_excel_col).value == "Yes":
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = tracked_fill
        elif "No" in gap_check_vals:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = gap_fill

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{header_row}"

    return rows, TRACKED_COL, FOAM_ANY_COL


def main():
    print("Step 1/5 — querying Item Master for FINISHED PRODUCT (FG) and INTERMEDIATE (SFG) codes...")
    conn = get_index()
    fg_master = dict(conn.execute(
        "SELECT code, desc FROM items WHERE item_type = 'FINISHED PRODUCT' AND status = 'ACTIVE'").fetchall())
    sfg_master = dict(conn.execute(
        "SELECT code, desc FROM items WHERE item_type = 'INTERMEDIATE' AND status = 'ACTIVE'").fetchall())
    print(f"  {len(fg_master):,} FINISHED PRODUCT codes, {len(sfg_master):,} INTERMEDIATE (SFG) codes\n")

    print("Step 2/5 — scanning all 4 monthly FG ledgers...")
    per_month_fg = {}
    for label, rel in zip(MONTH_LABELS, FG_FILES):
        path = os.path.join(RM_ROOT, rel)
        if not os.path.exists(path):
            print(f"  [{label}] skip — file not found")
            per_month_fg[label] = {}
            continue
        print(f"  [{label}] scanning {os.path.basename(path)}...")
        per_month_fg[label] = scan_ledger_codes(path, 'FG')

    print("\nStep 3/5 — scanning all 4 monthly SFG ledgers...")
    per_month_sfg = {}
    for label, rel in zip(MONTH_LABELS, SFG_FILES):
        path = os.path.join(RM_ROOT, rel)
        if not os.path.exists(path):
            print(f"  [{label}] skip — file not found")
            per_month_sfg[label] = {}
            continue
        print(f"  [{label}] scanning {os.path.basename(path)}...")
        per_month_sfg[label] = scan_ledger_codes(path, 'SFG')

    print("\nStep 4/5 — scanning all 4 monthly Foam Item Codes - Rate workbooks...")
    per_month_foam = {}
    for label, rel in zip(MONTH_LABELS, FOAM_FILES):
        path = os.path.join(RM_ROOT, rel)
        if not os.path.exists(path):
            print(f"  [{label}] skip — foam file not found")
            per_month_foam[label] = {}
            continue
        codes = scan_foam_codes(path)
        per_month_foam[label] = codes
        print(f"  [{label}] {os.path.basename(path)}: {len(codes):,} distinct foam codes")

    print("\nStep 5/5 — combining into FG Coverage + SFG Coverage sheets...")
    known = {s['item_code']: s for s in SKUS if s.get('item_code')}
    referenced_sfg = bom_store.all_referenced_sfg_codes()

    all_fg_codes = set(fg_master)
    for m in list(per_month_fg.values()) + list(per_month_foam.values()):
        all_fg_codes.update(m)

    all_sfg_codes = set(sfg_master) | referenced_sfg
    for m in list(per_month_sfg.values()) + list(per_month_foam.values()):
        all_sfg_codes.update(m)

    def fg_tracked_row(code):
        tracked = known.get(code)
        rm_snap = costing_store.latest(code) if tracked else None
        bom_months = bom_store.months_matched(code) if tracked else []
        return [
            "Yes" if tracked else "No",
            tracked['product'] if tracked else "",
            round(rm_snap[1], 2) if rm_snap else None,
            ", ".join(bom_months) if bom_months else "",
        ]

    def sfg_tracked_row(code):
        return ["Yes" if code in referenced_sfg else "No"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fg_rows, fg_tracked_col, fg_foam_col = build_coverage_sheet(
        wb, "FG Coverage", "FG Code", all_fg_codes, fg_master, "Item Master (FG)",
        per_month_fg, per_month_foam,
        ["Tracked in Tool?", "Matched Product", "RM Cost (Rs.)", "BOM Months Matched"],
        fg_tracked_row)

    sfg_rows, sfg_tracked_col, sfg_foam_col = build_coverage_sheet(
        wb, "SFG Coverage", "SFG Code", all_sfg_codes, sfg_master, "Item Master (SFG)",
        per_month_sfg, per_month_foam,
        ["Used in Tool's BOM Data?"],
        sfg_tracked_row)

    print("Writing workbook (this can take a minute for sheets this size)...")
    wb.save(OUT_PATH)

    fg_tracked_count = sum(1 for r in fg_rows if r[fg_tracked_col] == "Yes")
    fg_in_all_months = sum(1 for r in fg_rows if all(v == "Yes" for v in r[3:7]))
    sfg_used_count = sum(1 for r in sfg_rows if r[sfg_tracked_col] == "Yes")
    sfg_in_all_months = sum(1 for r in sfg_rows if all(v == "Yes" for v in r[3:7]))

    print(f"\nWrote {OUT_PATH}")
    print(f"\nFG Coverage: {len(fg_rows):,} distinct codes | {fg_tracked_count} tracked in tool | "
          f"{fg_in_all_months:,} present in all 4 months' FG ledgers")
    print(f"SFG Coverage: {len(sfg_rows):,} distinct codes | {sfg_used_count} used in tool's extracted BOMs | "
          f"{sfg_in_all_months:,} present in all 4 months' SFG ledgers")
    for lbl in MONTH_LABELS:
        print(f"  {lbl}: {len(per_month_fg[lbl]):,} FG-ledger codes, {len(per_month_sfg[lbl]):,} SFG-ledger codes, "
              f"{len(per_month_foam[lbl]):,} foam codes")


if __name__ == "__main__":
    main()
