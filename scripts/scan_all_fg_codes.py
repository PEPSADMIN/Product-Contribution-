"""
scan_all_fg_codes.py — full audit of every finished-product code in the
Ramco Item Master, cross-referenced against what this tool already tracks
in sku_master.py, for the user to hand to their team as a "here's what we
have vs. what Ramco has" validation sheet.

Uses the Item Master (Item Master.csv, 201,688 rows) as the reference —
not the monthly FG ledgers — because it's the actual master catalog of
every code Ramco has ever created (121,104 rows with Item Type =
"FINISHED PRODUCT"), all marked ACTIVE. A ledger only shows codes that
had a BOM transaction in that specific month, so it under- and
inconsistently represents the true catalog; Item Master is the
authoritative source and, being pre-indexed by item_master.py, this reads
in seconds instead of streaming hundreds of MB of monthly ledger files.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from item_master import get_index
from sku_master import SKUS

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "validation_exports", "All_FG_Codes_Validation.xlsx")


def main():
    conn = get_index()
    print("Querying Item Master for all FINISHED PRODUCT / ACTIVE codes...")
    fg_rows = conn.execute(
        "SELECT code, desc, uom, rate FROM items WHERE item_type = 'FINISHED PRODUCT' AND status = 'ACTIVE'"
    ).fetchall()
    print(f"  {len(fg_rows):,} finished-product codes found in Item Master")

    known_codes = {s['item_code']: s for s in SKUS if s.get('item_code')}

    rows = []
    for code, desc, uom, rate in fg_rows:
        known = known_codes.get(code)
        rows.append([
            code, desc, uom, round(rate, 2) if rate is not None else None,
            "Yes" if known else "No",
            known['product'] if known else "",
            (known.get('brand', '') if known else ""),
        ])
    rows.sort(key=lambda r: (r[4] != "Yes", r[0]))  # tracked codes first, then alphabetical

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All FG Codes"

    ws.append([f"All Finished-Product codes in the Ramco Item Master — {len(rows):,} distinct codes"])
    ws.append([])
    headers = ["FG Code", "Description", "UOM", "Standard Cost (Rs.)",
               "Tracked in Tool?", "Matched Product", "Matched Brand"]
    ws.append(headers)
    header_row = 3
    for r in rows:
        ws.append(r)

    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [26, 46, 10, 16, 15, 26, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    not_tracked_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    for r in range(header_row + 1, header_row + 1 + len(rows)):
        if ws.cell(row=r, column=5).value == "No":
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = not_tracked_fill

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

    print("Writing workbook...")
    wb.save(OUT_PATH)

    tracked = sum(1 for r in rows if r[4] == "Yes")
    print(f"\nWrote {OUT_PATH}")
    print(f"  {len(rows):,} finished-product codes in Item Master")
    print(f"  {tracked}/{len(known_codes)} of sku_master.py's known item codes were found among them")
    print(f"  {len(rows) - tracked:,} Item Master codes are NOT currently tracked as a SKU in this tool")

    missing_known = sorted(set(known_codes) - {r[0] for r in rows if r[4] == "Yes"})
    if missing_known:
        print(f"\n  {len(missing_known)} of sku_master.py's own codes were NOT found in Item Master "
              f"(check for typos or discontinued codes):")
        for code in missing_known:
            print(f"    {code} ({known_codes[code]['product']})")


if __name__ == "__main__":
    main()
