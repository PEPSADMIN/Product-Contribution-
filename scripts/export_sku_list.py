"""
export_sku_list.py — one Excel sheet listing every SKU (Brand/Product/Item
Code) plus its RM-cost and BOM-extraction coverage, for the user to
manually validate against their own records.

Read-only report: pulls from sku_master.py (the SKU list itself),
costing_store.py (latest RM cost snapshot per item_code), and bom_store.py
(which months' FG ledger actually matched each item_code) — doesn't touch
or recompute anything.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from sku_master import SKUS
import costing_store
import bom_store

BRAND_LABEL = {
    "peps": "Peps Spring",
    "cirrus": "Cirrus Foam",
    "italiano": "Italiano",
    "accessories": "Accessories",
}

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "validation_exports", "SKU_List_for_Validation.xlsx")

HEADERS = [
    "Brand", "Product", "Item Code", "Has Item Code?",
    "RM Cost (Rs.)", "RM Cost Source Month",
    "BOM Extracted?", "BOM Months Matched",
]


def main():
    rows = []
    for s in SKUS:
        code = s.get("item_code") or ""
        brand_label = BRAND_LABEL.get(s["brand"], s["brand"])

        rm_snap = costing_store.latest(code) if code else None
        rm_cost = rm_snap[1] if rm_snap else s.get("rm_cost")
        rm_month = rm_snap[0] if rm_snap else ""

        months = bom_store.months_matched(code) if code else []
        bom_extracted = "Yes" if months else "No"

        rows.append([
            brand_label, s["product"], code or "(none)",
            "Yes" if code else "No",
            round(rm_cost, 2) if rm_cost is not None else None,
            rm_month or "(not ledger-verified)",
            bom_extracted,
            ", ".join(months) if months else "(never matched)",
        ])

    rows.sort(key=lambda r: (r[0], r[1]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKU List"

    ws.append([f"Product Contribution Tool — SKU List for Validation ({len(rows)} SKUs)"])
    ws.append([])
    ws.append(HEADERS)
    header_row = 3
    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [16, 30, 26, 15, 14, 18, 15, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    no_code_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    no_bom_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    for r in range(header_row + 1, header_row + 1 + len(rows)):
        if ws.cell(row=r, column=4).value == "No":
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = no_code_fill
        elif ws.cell(row=r, column=7).value == "No":
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = no_bom_fill

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row}"

    wb.save(OUT_PATH)

    with_code = sum(1 for r in rows if r[3] == "Yes")
    with_bom = sum(1 for r in rows if r[6] == "Yes")
    print(f"Wrote {OUT_PATH}")
    print(f"  {len(rows)} total SKUs")
    print(f"  {with_code}/{len(rows)} have a Ramco item code")
    print(f"  {with_bom}/{len(rows)} have an extracted BOM (yellow = no item code, red = item code but never matched in the ledger)")


if __name__ == "__main__":
    main()
